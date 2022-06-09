from calendar import WEDNESDAY
import random
import time
from time import sleep
from datetime import datetime
import requests
from random import randint
import selenium.common.exceptions
from pygame import mixer
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import openpyxl

# F1 for students
# B2 for Tourists
visa_type = "F1"

# one for single person
# two for couples
number = "one"

# Chrome Webdriver
s = Service('../../Desktop/chromedriver_win32/chromedriver.exe')
driver = webdriver.Chrome(service=s)

# Open Excel file for saving data
book = openpyxl.Workbook()
sheet1 = book.active

# open temporary email website
url = "https://mail.tm/en/"
driver.get(url)

try:
    # 10 is the maximum time to wait
    element = WebDriverWait(driver, 10).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#address"))
    )
except TimeoutException:
    print("Copy Email Time out")


for j in range(100):
    j += 1
    try:
        mail = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "address"))).get_attribute('value')
        
        if "@" in mail:
            print(mail)
        else:
            sleep(5)
            driver.delete_all_cookies()
            driver.refresh()
            continue
    except:
        driver.refresh()

    # Open a new window
    driver.execute_script("window.open('');")

    # Switch to the new window
    driver.switch_to.window(driver.window_handles[1])
    
    # Let's Create account
    driver.get("https://ais.usvisa-info.com/en-ae/niv/signup")

    # Fake Account Info
    name = "Zahra"
    surname = "Ahmadi"
    password = "123456789"
    mobile = "9111699002"
    Name = driver.find_element(By.ID, "user_first_name").send_keys(name)
    Surname = driver.find_element(By.ID, "user_last_name").send_keys(surname)
    email1 = driver.find_element(By.ID, "user_email").send_keys(mail)
    email2 = driver.find_element(By.ID, "user_email_confirmation").send_keys(mail)
    pass1 = driver.find_element(By.ID, "user_password").send_keys(password)
    pass2 = driver.find_element(By.ID, "user_password_confirmation").send_keys(password)
    check_box = driver.find_element(By.CSS_SELECTOR,
                                    "#new_user > div.input.row.boolean.optional.user_mobile_alerts > div > label > div").click()
    sleep(1)
    Country_code = driver.find_element(By.ID, "user_mobile_country_code")
    Country_code_IR = Select(Country_code)
    Country_code_IR.select_by_value("IR")
    Mobile_Phone = driver.find_element(By.ID, "user_mobile_phone").send_keys(mobile)
    check_box = driver.find_element(By.CSS_SELECTOR, "#new_user > div:nth-child(7) > div > label > div").click()
    Create_Account = driver.find_element(By.NAME, "commit").click()

    try:
        # 10 is the maximum time to wait
        element = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "(//a[normalize-space()='update the email address'])[1]"))
        )
    except TimeoutException:
        driver.close()
        print("Sign up failed")
        driver.switch_to.window(driver.window_handles[0])
        sleep(5)
        driver.find_element(By.CSS_SELECTOR, "img[alt='Avatar']").click()
        sleep(1)
        driver.find_element(By.XPATH, "//a[normalize-space()='Sign out']").click()
        sleep(1)
        sleep(10)
        continue

    # close the active tab
    driver.close()

    # Switch back to the first tab
    driver.switch_to.window(driver.window_handles[0])

    for q in range(10):
        sleep(5)
        
        # Click on Confirmation link
        try:
            # 10 is the maximum time to wait
            element = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//div[normalize-space()='Confirmation instructions']"))).click()

        except:
            print("confirm error")
            driver.refresh()
            continue

        try:
            sleep(3)
            iframe = driver.find_element(By.ID, "iFrameResizer0")
            driver.switch_to.frame(iframe)
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.XPATH, "//a[contains(text(),'https://ais.usvisa-info.com/en-ae/niv/users/confir')]"))).click()
            break
        except selenium.common.exceptions.NoSuchElementException:
            driver.find_element(By. XPATH, "//a[normalize-space()='Inbox']").click()
            print("inbox")
            continue

    # Your Account is Activated
    driver.switch_to.window(driver.window_handles[1])
    driver.get("https://ais.usvisa-info.com/en-ae/niv/account")

    # Country / Authority that issued Passport
    passport = driver.find_element(By.ID, "applicant_passport_country_code")
    passport_IR = Select(passport)
    passport_IR.select_by_value("IR")

    # Country of Birth
    birth = driver.find_element(By.ID, "applicant_birth_country_code")
    birth_IR = Select(birth)
    birth_IR.select_by_value("IR")

    # Country of Permanent Residence
    residency = driver.find_element(By.ID, "applicant_permanent_residency_country_code")
    residency_IR = Select(residency)
    residency_IR.select_by_value("ir")

    # Passport Number
    passport_number = "T190517"
    driver.find_element(By.ID, "applicant_passport_number").send_keys(passport_number)

    # DS-160 Number
    DS160 = "AA00AEUGS4"
    driver.find_element(By.ID, "applicant_ds160_number").send_keys(DS160)

    # Visa Class
    if visa_type == "F1":
        # F1 (Student)
        visa = driver.find_element(By.ID, "applicant_visa_class_id")
        visa_usa = Select(visa)
        visa_usa.select_by_value("11")
    elif visa_type == "B2":
        # B2 (Tourist)
        visa = driver.find_element(By.ID, "applicant_visa_class_id")
        visa_usa = Select(visa)
        visa_usa.select_by_value("3")

    # Date of Birth
    birth_date = driver.find_element(By.ID, "applicant_date_of_birth_3i")
    b_date = Select(birth_date)
    b_date.select_by_value("5")

    birth_month = driver.find_element(By.ID, "applicant_date_of_birth_2i")
    b_month = Select(birth_month)
    b_month.select_by_value("9")

    birth_year = driver.find_element(By.ID, "applicant_date_of_birth_1i")
    b_year = Select(birth_year)
    b_year.select_by_value("1975")

    # Gender
    gender = driver.find_element(By.ID, "applicant_gender")
    a_gender = Select(gender)
    a_gender.select_by_value("F")

    # Primary Phone
    driver.find_element(By.ID, "applicant_phone1").send_keys(mobile)

    # Mobile country code & Phone Number --- div[class='icheckbox icheck-item icheck[h8pz3] icheck-area-20 hover']
    driver.find_element(By.CSS_SELECTOR,
                        "#applicant-creation-form > div > div > div:nth-child(13) > div.medium-4.small-12.columns.end.radio-checkbox-group > div > label > div").click()
    sleep(0.5)
    Country_code = driver.find_element(By.ID, "applicant_mobile_country_code")
    Country_code_IR = Select(Country_code)
    Country_code_IR.select_by_value("IR")

    # Mobile Phone
    driver.find_element(By.ID, "applicant_mobile_phone").send_keys(mobile)

    # Were you previously issued a visa to enter the United States? No
    driver.find_element(By.CSS_SELECTOR, "label[for='applicant_is_a_renewal_false']").click()

    # Is the applicant traveling from another country to apply for a visa in Armenia? Yes
    driver.find_element(By.CSS_SELECTOR, "label[for='applicant_traveling_to_apply_true']").click()

    # Create Applicant
    driver.find_element(By.CSS_SELECTOR, "input[value='Create Applicant']").click()

    # Confirm Applicant and Visa Information
    driver.find_element(By.CSS_SELECTOR, "input[value='Yes']").click()

    if visa_type == "B2":
        # Check all statements that apply
        driver.find_element(By. CSS_SELECTOR, "label[for='none_apply'] div[class='text']").click()
        driver.find_element(By. XPATH, "//input[@name='commit']").click()
    elif visa_type == "F1":
        pass

    # Do you want to add another applicant? No
    if number == "one":
        driver.find_element(By.CSS_SELECTOR, "div[class='off-canvas-content'] a:nth-child(2)").click()
    elif number == "two":
        driver.find_element(By. CSS_SELECTOR, "main[id='main'] a:nth-child(1)").click()

        # Second person information
        first = "Ali"
        last = "Jafari"
        driver.find_element(By.ID, "applicant_first_name").send_keys(first)
        driver.find_element(By.ID, "applicant_last_name").send_keys(last)

        # Country / Authority that issued Passport
        passport = driver.find_element(By.ID, "applicant_passport_country_code")
        passport_IR = Select(passport)
        passport_IR.select_by_value("IR")

        # Country of Birth
        birth = driver.find_element(By.ID, "applicant_birth_country_code")
        birth_IR = Select(birth)
        birth_IR.select_by_value("IR")

        # Country of Permanent Residence
        residency = driver.find_element(By.ID, "applicant_permanent_residency_country_code")
        residency_IR = Select(residency)
        residency_IR.select_by_value("ir")

        # Passport Number
        passport_number = "U111617"
        driver.find_element(By.ID, "applicant_passport_number").send_keys(passport_number)

        # DS-160 Number
        DS160 = "AA00AEPLS4"
        driver.find_element(By.ID, "applicant_ds160_number").send_keys(DS160)

        # Visa Class
        if visa_type == "F1":
            # F1 (Student)
            visa = driver.find_element(By.ID, "applicant_visa_class_id")
            visa_usa = Select(visa)
            visa_usa.select_by_value("11")
        elif visa_type == "B2":
            # B2 (Tourist)
            visa = driver.find_element(By.ID, "applicant_visa_class_id")
            visa_usa = Select(visa)
            visa_usa.select_by_value("3")

        # Date of Birth
        birth_date = driver.find_element(By.ID, "applicant_date_of_birth_3i")
        b_date = Select(birth_date)
        b_date.select_by_value("27")

        birth_month = driver.find_element(By.ID, "applicant_date_of_birth_2i")
        b_month = Select(birth_month)
        b_month.select_by_value("2")

        birth_year = driver.find_element(By.ID, "applicant_date_of_birth_1i")
        b_year = Select(birth_year)
        b_year.select_by_value("1980")

        # Gender
        gender = driver.find_element(By.ID, "applicant_gender")
        a_gender = Select(gender)
        a_gender.select_by_value("M")

        # Primary Phone
        driver.find_element(By.ID, "applicant_phone1").send_keys(mobile)

        # Mobile country code & Phone Number --- div[class='icheckbox icheck-item icheck[h8pz3] icheck-area-20 hover']
        driver.find_element(By.CSS_SELECTOR,
                            "#applicant-creation-form > div > div > div:nth-child(13) > div.medium-4.small-12.columns.end.radio-checkbox-group > div > label > div").click()
        sleep(0.5)
        Country_code = driver.find_element(By.ID, "applicant_mobile_country_code")
        Country_code_IR = Select(Country_code)
        Country_code_IR.select_by_value("IR")

        # Mobile Phone
        mobile = "09178993261"
        driver.find_element(By.ID, "applicant_mobile_phone").send_keys(mobile)

        # Were you previously issued a visa to enter the United States? No
        # CSS - #applicant-creation-form > div > div > div:nth-child(18) > div > div > div > div:nth-child(3) > label > div
        driver.find_element(By.CSS_SELECTOR, "label[for='applicant_is_a_renewal_false']").click()

        # Is the applicant traveling from another country to apply for a visa in Armenia? Yes
        # CSS -- #applicant-creation-form > div > div > div:nth-child(20) > div > div > div > div:nth-child(3) > label > div
        driver.find_element(By.CSS_SELECTOR, "label[for='applicant_traveling_to_apply_true']").click()

        # Create Applicant
        driver.find_element(By.CSS_SELECTOR, "input[value='Create Applicant']").click()

        # Does this visa type match your travel?
        driver.find_element(By. CSS_SELECTOR, "input[value='Yes']").click()

        if visa_type == "B2":
            # Check all statements that apply
            driver.find_element(By.CSS_SELECTOR, "label[for='none_apply'] div[class='text']").click()
            driver.find_element(By.XPATH, "//input[@name='commit']").click()
        elif visa_type == "F1":
            pass

        driver.find_element(By.CSS_SELECTOR, "div[class='off-canvas-content'] a:nth-child(2)").click()

    # Document courier pick up location
    Courier = driver.find_element(By.ID, "group_delivery_address_id")
    Courier = Select(Courier)
    Courier.select_by_value("3039551")
    driver.find_element(By.CSS_SELECTOR, "input[value='Continue']").click()

    try:
        element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable(
            (By.CSS_SELECTOR, "td[class='text-right']")))
    except:
        driver.refresh()

    # First Available Appointment
    Date = driver.find_element(By.CSS_SELECTOR, "td[class='text-right']").text
    now = datetime.now().strftime("%H:%M:%S")

    # Get current link
    cur_url = driver.current_url

    # Save fake account login and link information
    sheet1.cell(row=j, column=1, value=mail)
    sheet1.cell(row=j, column=2, value=password)
    sheet1.cell(row=j, column=3, value=cur_url)
    book.save('Fake.xlsx')

    # Signout 
    driver.get("https://ais.usvisa-info.com/en-ae/niv/users/sign_out")
    driver.delete_all_cookies()

    # back to temporary email generator 
    driver.switch_to.window(driver.window_handles[0])
    sleep(1)
    driver.find_element(By. XPATH, "//button[@id='logout']//*[name()='svg']").click()
    sleep(1)
    driver.delete_all_cookies()
    driver.close()
    driver.switch_to.window(driver.window_handles[0])
    url = "https://mail.tm/en/"
    driver.get(url)