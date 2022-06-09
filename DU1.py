import random
import openpyxl
import pause
from time import sleep
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import DU2
sleep(5)

#s = Service('../../Desktop/chromedriver_win32/chromedriver.exe')
#driver = webdriver.Chrome(service=s)
driver = webdriver.Chrome()

# Sign out function
def sign_out():
    driver.get("https://ais.usvisa-info.com/en-ae/niv/users/sign_out")
    driver.delete_all_cookies()
    
# Login Function            
def login():
    global TEXT
    
    for i in range(3):
        # select randomly one email from fake accounts to login
        WB = openpyxl.load_workbook("fake.xlsx")
        Sheet = WB.worksheets[0]
        row = Sheet.max_row
        r = random.randrange(1, row)
        user = Sheet.cell(r, 1).value
        password = Sheet.cell(r, 2).value
        link = Sheet.cell(r, 3).value
        
        print("Email address: " + str(user))
        driver.get(link)

        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR,
                        "button[class='ui-button ui-corner-all ui-widget']"))).click()
        except:
            TEXT = "Dubai - Pop Up failed"
            print(TEXT)
            sleep(5)
            driver.get("https://ais.usvisa-info.com/en-ae/niv/users/sign_out")
            driver.delete_all_cookies()
            continue
        break

    driver.find_element(By.ID, "user_email").send_keys(user)
    driver.find_element(By.ID, "user_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR,
                        "#new_user > div.radio-checkbox-group.margin-top-30 > label > div").click()
    driver.find_element(By.CSS_SELECTOR, "input[value='Sign In']").click()

login()

def visa():
    global now
    global Date
    
    for i in range(2):
        
        # Read first availabe appointment
        driver.refresh()
        
        try:
            Date = WebDriverWait(driver, 5).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR, "td[class='text-right']"))).text
        except:
            sleep(0.25)
            DU2.visa()
            driver.delete_all_cookies()
            login()
            sleep(5)
            return

        now = datetime.now().strftime("%H:%M:%S")

        if Date != "No Appointments Available":
            Date = str(Date)
            Date = Date.replace(',', "")
            date = open("FAD.txt", 'w')
            date.write(Date)
            date.close()
            
            # day of month
            day = int(Date[0:2])

            output = now + " >>> " + Date
            print(output)
            
            # check Month
            if "April 2022" in Date or "May 2022" in Date or "June 2022" in Date or "July 2022" in Date:
                DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.app.visa()
            else:
                DU2.visa()
        else:
            output = now + " >>> " + Date
            print(output)
            DU2.visa()
        sleep(1)    

while True:
    for k in range(45):
        if k == 44:
            print("-----------------------------------------------------")
            print("After 4 hours, Lets sign out and change Accounts :-) ")
            print("-----------------------------------------------------")
            sleep(5)
            
            sign_out()
            sleep(1)
            login()
            sleep(1)
            DU2.sign_out()
            sleep(1)
            DU2.login()
            sleep(1)
            DU2.DU3.sign_out()
            sleep(1)
            DU2.DU3.login()
            sleep(1)
            DU2.DU3.DU4.sign_out()
            sleep(1)
            DU2.DU3.DU4.login()
            sleep(1)
            DU2.DU3.DU4.DU5.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.login()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.app.sign_out()
            sleep(1)
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.app.login()
        else:
            pass
            
        tm = datetime.now()
        tm = tm - timedelta(minutes=tm.minute % 5,
                            seconds=tm.second,
                            microseconds=tm.microsecond)

        year, month, day, mn, hr = tm.year, tm.month, tm.day, tm.minute, tm.hour

        if mn == 55:
            if hr == 23:
                sleep(300)
                mn = 5
                hr = 0
            else:
                mn = 0
                hr = hr + 1
        else:
            mn = mn + 5

        if mn == 20 or mn == 50:
            DU2.DU3.DU4.DU5.DU6.DU7.DU8.DU9.DU10.app.refresh()
        else:
            pass
        
        sec = 24
        print(
            "Next run at: " + "2022" + "-" + str(month) + "-" + str(day) + ", " + str(hr) + ":" + str(mn) + ":" + str(
                sec))
        pause.until(datetime(year, month, day, hr, mn, sec))
        visa()