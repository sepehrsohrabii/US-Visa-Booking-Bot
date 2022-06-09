import random
import openpyxl
from time import sleep
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.select import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import DU10
sleep(5)

#s = Service('../../Desktop/chromedriver_win32/chromedriver.exe')
driver = webdriver.Chrome()

def sign_out():
    driver.get("https://ais.usvisa-info.com/en-ae/niv/users/sign_out")
    driver.delete_all_cookies()
    
            
def login():
    global TEXT
    
    for i in range(3):
        WB = openpyxl.load_workbook("fake.xlsx")
        Sheet = WB.worksheets[0]
        row = Sheet.max_row
        r = random.randrange(1, row)
        user = Sheet.cell(r, 1).value
        password = Sheet.cell(r, 2).value
        link = Sheet.cell(r, 3).value
        
        print("Changing Email address to: " + str(user))
        driver.get(link)

        try:
            WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable(
                    (By.CSS_SELECTOR,
                        "button[class='ui-button ui-corner-all ui-widget']"))).click()
        except:
            TEXT = "Dubai - mail Pop Up failed, STOP WORKING"
            print(TEXT)
            
            sleep(15)
            driver.get("https://ais.usvisa-info.com/en-ae/niv/users/sign_out")
            driver.delete_all_cookies()
            continue
        break


    driver.find_element(By.ID, "user_email").send_keys(user)
    driver.find_element(By.ID, "user_password").send_keys(password)
    driver.find_element(By.CSS_SELECTOR,
                        "#new_user > div.radio-checkbox-group.margin-top-30 > label > div").click()
    driver.find_element(By.CSS_SELECTOR, "input[value='Sign In']").click()

login()

def visa():
    global now
    global Date
    driver.refresh()
    
    try:
        Date = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located(
                (By.CSS_SELECTOR, "td[class='text-right']"))).text
    except:
        DU10.visa()
        driver.delete_all_cookies()
        login()
        sleep(5)
        return

    now = datetime.now().strftime("%H:%M:%S")

    if Date != "No Appointments Available":
        Date = str(Date)
        Date = Date.replace(',', "")
        day = int(Date[0:2])
        
        output = now + " >>> " + Date
        print(output)
        
        date = open('FAD.txt', 'r')
        date = date.read()
        
        if date != Date:
            FAD = open('FAD.txt', 'w')
            FAD.write(Date)
            FAD.close()
            
            if "April 2022" in Date or "May 2022" in Date or "June 2022" in Date or "July 2022" in Date:
                DU10.app.visa()
            else:
                DU10.visa()
        else:        
            DU10.visa()     
             
        return
    else:
        output = now + " >>> " + Date
        print(output)
        DU10.visa()